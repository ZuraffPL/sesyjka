# -*- coding: utf-8 -*-
"""
Moduł zarządzania ścieżkami i migracjami baz danych
Zapewnia kompatybilność wsteczną przy aktualizacjach aplikacji
"""

import os
import sqlite3
import shutil
import zipfile
import tempfile
from pathlib import Path
from typing import Optional, Tuple, List
from datetime import datetime

# Wersja schematu bazy danych
CURRENT_DB_VERSION = 1

# ── Nazwy plików baz danych ───────────────────────────────────────────────────
_DB_FILES: List[str] = ["systemy_rpg.db", "sesje_rpg.db", "gracze.db", "wydawcy.db"]

# ── Tryb gościa — globalny katalog override ───────────────────────────────────
_guest_db_dir: Optional[Path] = None


def set_guest_db_dir(path: Optional[Path]) -> None:
    """Ustawia katalog baz danych gościa (None = powrót do własnych danych)."""
    global _guest_db_dir
    _guest_db_dir = path


def is_guest_mode() -> bool:
    """Zwraca True jeśli aplikacja działa na bazach gościa (tylko odczyt)."""
    return _guest_db_dir is not None


def get_app_data_dir() -> Path:
    """
    Pobiera katalog danych aplikacji w folderze użytkownika.
    Struktura: C:/Users/{username}/AppData/Local/Sesyjka/

    Returns:
        Path: Ścieżka do katalogu danych aplikacji
    """
    if os.name == 'nt':  # Windows
        app_data = os.getenv('LOCALAPPDATA', os.path.expanduser('~'))
        app_dir = Path(app_data) / 'Sesyjka'
    else:  # Linux/Mac
        app_dir = Path.home() / '.sesyjka'

    # Utwórz katalog jeśli nie istnieje
    app_dir.mkdir(parents=True, exist_ok=True)
    return app_dir


def get_db_path(db_name: str) -> str:
    """
    Pobiera pełną ścieżkę do pliku bazy danych.
    W trybie gościa zwraca ścieżkę z katalogu gościa.

    Args:
        db_name: Nazwa pliku bazy danych (np. 'systemy_rpg.db')

    Returns:
        str: Pełna ścieżka do pliku bazy danych
    """
    if _guest_db_dir is not None:
        return str(_guest_db_dir / db_name)
    app_dir = get_app_data_dir()
    db_path = app_dir / db_name
    return str(db_path)


def get_own_db_path(db_name: str) -> str:
    """
    Zawsze zwraca ścieżkę do własnych baz danych, ignorując tryb gościa.
    Używać przy eksporcie i tworzeniu backupów.

    Args:
        db_name: Nazwa pliku bazy danych

    Returns:
        str: Pełna ścieżka do własnej bazy danych
    """
    return str(get_app_data_dir() / db_name)


def migrate_old_databases() -> None:
    """
    Migruje stare bazy danych z katalogu aplikacji do folderu użytkownika.
    Tworzy kopie zapasowe podczas migracji.
    """
    old_db_files = ["systemy_rpg.db", "sesje_rpg.db", "gracze.db", "wydawcy.db"]

    app_dir = get_app_data_dir()
    current_dir = Path.cwd()

    for db_file in old_db_files:
        old_path = current_dir / db_file
        new_path = app_dir / db_file

        # Jeśli stara baza istnieje w katalogu aplikacji i nie istnieje w nowej lokalizacji
        if old_path.exists() and not new_path.exists():
            try:
                # Skopiuj bazę danych do nowej lokalizacji
                shutil.copy2(old_path, new_path)
                print(f"✓ Zmigrowano {db_file} do {app_dir}")

                # Opcjonalnie: Utwórz backup starej bazy
                backup_name = f"{db_file}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                backup_path = current_dir / backup_name
                shutil.copy2(old_path, backup_path)
                print(f"✓ Utworzono backup: {backup_name}")

            except Exception as e:
                print(f"⚠ Błąd podczas migracji {db_file}: {e}")


def get_db_version(db_path: str) -> int:
    """
    Pobiera wersję schematu bazy danych.

    Args:
        db_path: Ścieżka do pliku bazy danych

    Returns:
        int: Numer wersji schematu (0 jeśli tabela nie istnieje)
    """
    if not os.path.exists(db_path):
        return 0

    try:
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA foreign_keys = ON")
            c = conn.cursor()
            c.execute(
                """
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name='db_version'
            """
            )

            if c.fetchone():
                c.execute("SELECT version FROM db_version LIMIT 1")
                result = c.fetchone()
                return result[0] if result else 0
            return 0
    except Exception:
        return 0


def set_db_version(db_path: str, version: int) -> None:
    """
    Ustawia wersję schematu bazy danych.

    Args:
        db_path: Ścieżka do pliku bazy danych
        version: Numer wersji do ustawienia
    """
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        c = conn.cursor()

        # Utwórz tabelę wersji jeśli nie istnieje
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS db_version (
                version INTEGER NOT NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """
        )

        # Usuń starą wersję i wstaw nową
        c.execute("DELETE FROM db_version")
        c.execute("INSERT INTO db_version (version) VALUES (?)", (version,))
        conn.commit()


def backup_database(db_path: str) -> Optional[str]:
    """
    Tworzy kopię zapasową bazy danych przed migracją.

    Args:
        db_path: Ścieżka do pliku bazy danych

    Returns:
        Optional[str]: Ścieżka do pliku backup lub None w przypadku błędu
    """
    if not os.path.exists(db_path):
        return None

    try:
        app_dir = get_app_data_dir()
        backups_dir = app_dir / 'backups'
        backups_dir.mkdir(exist_ok=True)

        db_name = Path(db_path).name
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"{db_name}.backup_{timestamp}"
        backup_path = backups_dir / backup_name

        shutil.copy2(db_path, backup_path)
        print(f"✓ Utworzono backup: {backup_path}")
        return str(backup_path)

    except Exception as e:
        print(f"⚠ Błąd podczas tworzenia backupu: {e}")
        return None


def migrate_database_schema(db_path: str, db_name: str) -> None:
    """
    Przeprowadza migrację schematu bazy danych do najnowszej wersji.

    Args:
        db_path: Ścieżka do pliku bazy danych
        db_name: Nazwa bazy danych (dla identyfikacji typu)
    """
    current_version = get_db_version(db_path)

    if current_version >= CURRENT_DB_VERSION:
        return  # Baza jest aktualna

    # Utwórz backup przed migracją
    backup_database(db_path)

    print(f"Migracja {db_name} z wersji {current_version} do {CURRENT_DB_VERSION}")

    # W przyszłości tutaj będą migracje dla różnych wersji
    # if current_version < 1:
    #     migrate_to_v1(db_path, db_name)
    # if current_version < 2:
    #     migrate_to_v2(db_path, db_name)

    # Ustaw nową wersję
    set_db_version(db_path, CURRENT_DB_VERSION)
    print(f"✓ Migracja {db_name} zakończona")


def initialize_app_databases() -> None:
    """
    Inicjalizuje wszystkie bazy danych aplikacji.
    Wykonuje migrację starych baz i sprawdza zgodność wersji.
    """
    print("=" * 60)
    print("Inicjalizacja baz danych Sesyjka")
    print("=" * 60)

    # Najpierw sprawdź czy są stare bazy do migracji
    migrate_old_databases()

    # Zainicjalizuj ścieżki do nowych baz
    db_configs = {
        'systemy_rpg.db': 'Systemy RPG',
        'sesje_rpg.db': 'Sesje RPG',
        'gracze.db': 'Gracze',
        'wydawcy.db': 'Wydawcy',
    }

    app_dir = get_app_data_dir()
    print(f"\nLokalizacja baz danych: {app_dir}")
    print("-" * 60)

    for db_file, db_label in db_configs.items():
        db_path = get_db_path(db_file)

        if os.path.exists(db_path):
            print(f"✓ {db_label}: Znaleziono istniejącą bazę")
            # Sprawdź wersję i ewentualnie zmigruj
            migrate_database_schema(db_path, db_file)
        else:
            print(f"→ {db_label}: Utworzenie nowej bazy")

    print("=" * 60)
    print("Inicjalizacja zakończona")
    print("=" * 60)
    print()


def ensure_app_icons() -> None:
    """
    Kopiuje ikony aplikacji do katalogu AppData przy każdym uruchomieniu.
    Dziła zarówno z kodu źródłowego jak i z pliku EXE (PyInstaller one-file).
    Dzięki temu ikona przycisku edycji jest dostępna na każdym komputerze.
    """
    import sys

    icons_dst = get_app_data_dir() / 'Icons'
    icons_dst.mkdir(parents=True, exist_ok=True)

    # Źródło ikon: PyInstaller _MEIPASS > katalog pliku
    if hasattr(sys, '_MEIPASS'):
        src = Path(getattr(sys, '_MEIPASS')) / 'Icons'
    else:
        src = Path(__file__).parent / 'Icons'

    if not src.exists():
        return

    for icon_file in src.glob('*.png'):
        dst = icons_dst / icon_file.name
        try:
            shutil.copy2(str(icon_file), str(dst))
        except Exception:
            pass


# ── Eksport baz danych ────────────────────────────────────────────────────────

def export_databases_excel(dest: Path) -> Path:
    """
    Eksportuje własne bazy danych do jednego pliku Excel (.xlsx).

    Każda tabela z każdej bazy danych staje się osobnym arkuszem.
    Arkusze są nazwane wg schematu: "NazwaBazy - nazwa_tabeli".

    Args:
        dest: Ścieżka docelowa — plik .xlsx.

    Returns:
        Path: Ścieżka do zapisanego pliku.

    Raises:
        ImportError: Gdy biblioteka openpyxl nie jest zainstalowana.
        ValueError: Gdy żadna baza danych nie istnieje.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError(
            "Eksport do Excela wymaga biblioteki openpyxl.\n"
            "Zainstaluj ją poleceniem: pip install openpyxl"
        )

    _DB_LABELS: dict[str, str] = {
        "systemy_rpg.db": "Systemy RPG",
        "sesje_rpg.db": "Sesje RPG",
        "gracze.db": "Gracze",
        "wydawcy.db": "Wydawcy",
    }

    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # usuń domyślny pusty arkusz

    own_dir = get_app_data_dir()

    for db_file in _DB_FILES:
        src = own_dir / db_file
        if not src.exists():
            continue
        label = _DB_LABELS.get(db_file, db_file.replace(".db", ""))

        conn = sqlite3.connect(src)
        conn.row_factory = sqlite3.Row
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
            )
            tables = [row[0] for row in cursor.fetchall() if not row[0].startswith("sqlite_")]

            for table in tables:
                # Excel ogranicza nazwy arkuszy do 31 znaków
                sheet_name = f"{label} - {table}"[:31]
                ws = wb.create_sheet(title=sheet_name)

                cursor.execute(f"SELECT * FROM [{table}] LIMIT 0")  # tylko nagłówki
                if cursor.description is None:
                    continue
                headers = [desc[0] for desc in cursor.description]

                # Nagłówki ze stylem
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align

                # Dane
                cursor.execute(f"SELECT * FROM [{table}]")
                for row_idx, row in enumerate(cursor.fetchall(), 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # Auto-szerokość kolumn (maks. 50 znaków)
                for col in ws.columns:
                    max_len = max(
                        (len(str(cell.value)) for cell in col if cell.value is not None),
                        default=8,
                    )
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        finally:
            conn.close()

    if not wb.sheetnames:
        raise ValueError("Brak baz danych do wyeksportowania.")

    wb.save(dest)
    return dest


def export_databases(dest: Path, fmt: str) -> Path:
    """
    Eksportuje własne bazy danych do pliku ZIP lub folderu.

    Args:
        dest: Ścieżka docelowa — plik .zip (fmt='zip') lub katalog (fmt='folder').
        fmt: 'zip' lub 'folder'.

    Returns:
        Path: Ścieżka do utworzonego pliku/folderu.

    Raises:
        ValueError: Gdy fmt jest nieznany.
        OSError: Gdy nie można zapisać pliku.
    """
    own_dir = get_app_data_dir()
    if fmt == 'zip':
        with zipfile.ZipFile(dest, 'w', zipfile.ZIP_DEFLATED) as zf:
            for db_file in _DB_FILES:
                src = own_dir / db_file
                if src.exists():
                    zf.write(src, db_file)
        return dest
    elif fmt == 'folder':
        dest.mkdir(parents=True, exist_ok=True)
        for db_file in _DB_FILES:
            src = own_dir / db_file
            if src.exists():
                shutil.copy2(src, dest / db_file)
        return dest
    else:
        raise ValueError(f"Nieznany format eksportu: {fmt!r}")


def prepare_import_source(source: Path) -> Tuple[Path, List[str]]:
    """
    Przygotowuje źródło importu — rozpakowuje ZIP lub waliduje folder.

    ZIP jest rozpakowywany do katalogu tymczasowego (%TEMP%\\sesyjka_import_*),
    który pozostaje do końca sesji (lub do jawnego usunięcia przez exit_guest_mode).

    Args:
        source: Ścieżka do pliku .zip lub folderu z bazami.

    Returns:
        Tuple[Path, List[str]]: (katalog_z_plikami_db, lista_znalezionych_plików_db)

    Raises:
        ValueError: Gdy źródło nie zawiera żadnych baz danych Sesyjki.
    """
    db_set = set(_DB_FILES)
    if source.suffix.lower() == '.zip':
        with zipfile.ZipFile(source) as zf:
            # Tylko pliki bezpośrednio w root ZIP (bez podkatalogów)
            names = {n for n in zf.namelist() if '/' not in n and '\\' not in n}
            found = sorted(names & db_set)
            if not found:
                raise ValueError("Plik ZIP nie zawiera żadnych baz danych Sesyjki (.db).")
            tmp = Path(tempfile.mkdtemp(prefix="sesyjka_import_"))
            for f in found:
                zf.extract(f, tmp)
        return tmp, found
    elif source.is_dir():
        found = sorted(f for f in _DB_FILES if (source / f).exists())
        if not found:
            raise ValueError("Folder nie zawiera żadnych baz danych Sesyjki (.db).")
        return source, found
    else:
        raise ValueError("Nieznany format — wybierz plik .zip lub folder z bazami Sesyjki.")


def replace_own_databases(source_dir: Path, db_files: List[str]) -> None:
    """
    Nadpisuje własne bazy danych plikami z source_dir.
    Przed każdym nadpisaniem tworzy backup w %LOCALAPPDATA%\\Sesyjka\\backups\\.

    Args:
        source_dir: Katalog źródłowy z plikami .db.
        db_files: Lista nazw plików .db do nadpisania.
    """
    own_dir = get_app_data_dir()
    for db_file in db_files:
        src = source_dir / db_file
        if not src.exists():
            continue
        dst = own_dir / db_file
        if dst.exists():
            backup_database(str(dst))
        shutil.copy2(src, dst)


# Eksportuj funkcje dla kompatybilności
__all__ = [
    'get_app_data_dir',
    'get_db_path',
    'get_own_db_path',
    'set_guest_db_dir',
    'is_guest_mode',
    'export_databases',
    'export_databases_excel',
    'prepare_import_source',
    'replace_own_databases',
    'migrate_old_databases',
    'initialize_app_databases',
    'ensure_app_icons',
    'backup_database',
    'CURRENT_DB_VERSION',
]
